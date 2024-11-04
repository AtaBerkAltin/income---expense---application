import tkinter as tk
from tkinter import messagebox, ttk
from tkcalendar import DateEntry
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

gelirler = []
giderler = []
tekrarlayan_giderler = []

# Döviz çevirme oranları yaklaşık düz hesap
exchange_rates = {
    "TL": 1,
    "USD": 35.0,  # Örnek kur: 1 USD = 35 TL
    "EUR": 37.0   # Örnek kur: 1 EUR = 37 TL
}

def para_birimine_cevir(miktar, birim, hedef_birim="TL"):
    """Para birimini hedef para birimine (TL) çevirir."""
    if birim == hedef_birim:
        return miktar
    tl_miktar = miktar * exchange_rates[birim]  # İlk önce TL'ye çevir
    return tl_miktar / exchange_rates[hedef_birim]  # TL'den hedef para birimine çevir

def gelir_ekle():
    try:
        gelir_turu = gelir_turu_entry.get()
        gelir_miktari = float(gelir_miktari_entry.get())
        para_birimi = gelir_para_birimi.get()
        tarih = gelir_tarihi_entry.get_date()
        saat = f"{gelir_saat_combobox.get()}:{gelir_dakika_combobox.get()}"
        gelirler.append({"Tür": gelir_turu, "Miktar": gelir_miktari, "Para Birimi": para_birimi, "Tarih": tarih, "Saat": saat})
        messagebox.showinfo("Başarılı", "Gelir başarıyla eklendi.")
    except ValueError:
        messagebox.showerror("Hata", "Geçerli bir miktar girin.")
    gelir_turu_entry.delete(0, tk.END)
    gelir_miktari_entry.delete(0, tk.END)

def gider_ekle():
    try:
        gider_turu = gider_turu_entry.get()
        gider_miktari = float(gider_miktari_entry.get())
        para_birimi = gider_para_birimi.get()
        tarih = gider_tarihi_entry.get_date()
        saat = f"{gider_saat_combobox.get()}:{gider_dakika_combobox.get()}"
        tekrarla = tekrarla_var.get()

        gider = {"Tür": gider_turu, "Miktar": gider_miktari, "Para Birimi": para_birimi, "Tarih": tarih, "Saat": saat}
        giderler.append(gider)
        
        # Tekrarlayan giderleri ayrı bir listeye kaydediyoruz
        if tekrarla:
            tekrarlayan_giderler.append(gider)

        messagebox.showinfo("Başarılı", "Gider başarıyla eklendi.")
    except ValueError:
        messagebox.showerror("Hata", "Geçerli bir miktar girin.")
    gider_turu_entry.delete(0, tk.END)
    gider_miktari_entry.delete(0, tk.END)

def hesapla():
    toplam_gelir = sum([para_birimine_cevir(gelir["Miktar"], gelir["Para Birimi"]) for gelir in gelirler])
    toplam_gider = sum([para_birimine_cevir(gider["Miktar"], gider["Para Birimi"]) for gider in giderler])
    net_gelir = toplam_gelir - toplam_gider

    sonuc_label.config(text=f"Toplam Gelir: {toplam_gelir:.2f} TL\nToplam Gider: {toplam_gider:.2f} TL\nNet Gelir: {net_gelir:.2f} TL")

def excel_kaydet():
    try:
        if not gelirler and not giderler:
            messagebox.showerror("Hata", "Kaydedilecek veri yok.")
            return

        # Tüm gelir ve giderleri TL'ye çevirerek yeni DataFrame'ler oluşturuyoruz
        gelirler_df = pd.DataFrame(
            [{"Tür": gelir["Tür"], "Miktar (TL)": para_birimine_cevir(gelir["Miktar"], gelir["Para Birimi"]), "Tarih": gelir["Tarih"], "Saat": gelir["Saat"]}
             for gelir in gelirler]
        )
        giderler_df = pd.DataFrame(
            [{"Tür": gider["Tür"], "Miktar (TL)": para_birimine_cevir(gider["Miktar"], gider["Para Birimi"]), "Tarih": gider["Tarih"], "Saat": gider["Saat"]}
             for gider in giderler]
        )

        # Yeni bir Excel dosyası oluştur
        wb = Workbook()
        
        # Gelirler sayfasını ekle ve biçimlendir
        if not gelirler_df.empty:
            ws_gelirler = wb.active
            ws_gelirler.title = "Gelirler"
            
            # Gelirler verilerini yaz
            for r in dataframe_to_rows(gelirler_df, index=False, header=True):
                ws_gelirler.append(r)
            
            # Hücrelerin genişliğini, yüksekliğini ve fontunu ayarla
            for row in ws_gelirler.iter_rows():
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(left=Side(style="thick"), right=Side(style="thick"), 
                                         top=Side(style="thick"), bottom=Side(style="thick"))
            for col in ws_gelirler.columns:
                col_letter = col[0].column_letter
                ws_gelirler.column_dimensions[col_letter].width = 20  # Yaklaşık 160px
            
            for row in range(1, ws_gelirler.max_row + 1):
                ws_gelirler.row_dimensions[row].height = 20  # Yaklaşık 80px

            # Tarih formatını ayarlayın
            for row in ws_gelirler.iter_rows(min_row=2, min_col=3, max_col=3):
                for cell in row:
                    cell.number_format = 'DD/MM/YYYY'  # Gün/Ay/Yıl formatı

        # Giderler sayfasını ekle ve biçimlendir
        if not giderler_df.empty:
            ws_giderler = wb.create_sheet(title="Giderler")
            
            # Giderler verilerini yaz
            for r in dataframe_to_rows(giderler_df, index=False, header=True):
                ws_giderler.append(r)
            
            # Hücrelerin genişliğini, yüksekliğini ve fontunu ayarla
            for row in ws_giderler.iter_rows():
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(left=Side(style="thick"), right=Side(style="thick"), 
                                         top=Side(style="thick"), bottom=Side(style="thick"))
            for col in ws_giderler.columns:
                col_letter = col[0].column_letter
                ws_giderler.column_dimensions[col_letter].width = 20  # Yaklaşık 160px
            
            for row in range(1, ws_giderler.max_row + 1):
                ws_giderler.row_dimensions[row].height = 20  # Yaklaşık 80px

            # Tarih formatını ayarlayın
            for row in ws_giderler.iter_rows(min_row=2, min_col=3, max_col=3):
                for cell in row:
                    cell.number_format = 'DD/MM/YYYY'  # Gün/Ay/Yıl formatı

        # Dosya yolunu masaüstüne kaydetmek için güncelleyebilirsiniz
        wb.save(r"C:\Users\user\OneDrive\Masaüstü\GelirGiderHesaplamasi.xlsx")
        messagebox.showinfo("Başarılı", "Veriler Excel dosyasına başarıyla kaydedildi.")
    except Exception as e:
        messagebox.showerror("Hata", f"Excel kaydedilirken bir hata oluştu: {e}")

def tekrarlayan_giderleri_kontrol_et():
    """Tekrarlayan giderleri kontrol edip, her ay başında gerekirse otomatik ekler."""
    bugun = datetime.now()
    for gider in tekrarlayan_giderler:
        son_ekleme_tarihi = gider["Tarih"]
        if (bugun - son_ekleme_tarihi).days >= 30:
            yeni_gider = gider.copy()
            yeni_gider["Tarih"] = bugun
            giderler.append(yeni_gider)

root = tk.Tk()
root.title("Gelir-Gider Takip Programı")
root.geometry("1200x800")  

# Gelir Bölümü
gelir_frame = tk.LabelFrame(root, text="Gelir Ekle", font=("Arial", 12, "bold"))
gelir_frame.pack(padx=10, pady=10, fill="both", expand=True)

gelir_turu_label = tk.Label(gelir_frame, text="Gelir Türü:", font=("Arial", 10))
gelir_turu_label.grid(row=0, column=0, padx=5, pady=5)
gelir_turu_entry = tk.Entry(gelir_frame, font=("Arial", 10), width=40)
gelir_turu_entry.grid(row=0, column=1, padx=5, pady=5)

gelir_miktari_label = tk.Label(gelir_frame, text="Gelir Miktarı:", font=("Arial", 10))
gelir_miktari_label.grid(row=1, column=0, padx=5, pady=5)
gelir_miktari_entry = tk.Entry(gelir_frame, font=("Arial", 10), width=40)
gelir_miktari_entry.grid(row=1, column=1, padx=5, pady=5)

gelir_para_birimi = ttk.Combobox(gelir_frame, values=["TL", "USD", "EUR"], font=("Arial", 10), state="readonly")
gelir_para_birimi.current(0)  # Varsayılan olarak TL seçili
gelir_para_birimi.grid(row=2, column=1, padx=5, pady=5)

gelir_para_birimi_label = tk.Label(gelir_frame, text="Para Birimi:", font=("Arial", 10))
gelir_para_birimi_label.grid(row=2, column=0, padx=5, pady=5)

gelir_tarihi_label = tk.Label(gelir_frame, text="Tarih:", font=("Arial", 10))
gelir_tarihi_label.grid(row=3, column=0, padx=5, pady=5)
gelir_tarihi_entry = DateEntry(gelir_frame, font=("Arial", 10), width=37, background='darkblue', foreground='white', borderwidth=2)
gelir_tarihi_entry.grid(row=3, column=1, padx=5, pady=5)

gelir_saat_combobox = ttk.Combobox(gelir_frame, values=[f"{i:02}" for i in range(24)], font=("Arial", 10), state="readonly")
gelir_saat_combobox.grid(row=4, column=1, padx=5, pady=5)
gelir_saat_combobox.current(0)

gelir_dakika_combobox = ttk.Combobox(gelir_frame, values=[f"{i:02}" for i in range(60)], font=("Arial", 10), state="readonly")
gelir_dakika_combobox.grid(row=5, column=1, padx=5, pady=5)
gelir_dakika_combobox.current(0)

gelir_saat_label = tk.Label(gelir_frame, text="Saat:", font=("Arial", 10))
gelir_saat_label.grid(row=4, column=0, padx=5, pady=5)

gelir_dakika_label = tk.Label(gelir_frame, text="Dakika:", font=("Arial", 10))
gelir_dakika_label.grid(row=5, column=0, padx=5, pady=5)

gelir_ekle_button = tk.Button(gelir_frame, text="Gelir Ekle", command=gelir_ekle, font=("Arial", 10), bg="lightgreen")
gelir_ekle_button.grid(row=6, columnspan=2, padx=5, pady=5, ipadx=10)

# Gider Bölümü
gider_frame = tk.LabelFrame(root, text="Gider Ekle", font=("Arial", 12, "bold"))
gider_frame.pack(padx=10, pady=10, fill="both", expand=True)

gider_turu_label = tk.Label(gider_frame, text="Gider Türü:", font=("Arial", 10))
gider_turu_label.grid(row=0, column=0, padx=5, pady=5)
gider_turu_entry = tk.Entry(gider_frame, font=("Arial", 10), width=40)
gider_turu_entry.grid(row=0, column=1, padx=5, pady=5)

gider_miktari_label = tk.Label(gider_frame, text="Gider Miktarı:", font=("Arial", 10))
gider_miktari_label.grid(row=1, column=0, padx=5, pady=5)
gider_miktari_entry = tk.Entry(gider_frame, font=("Arial", 10), width=40)
gider_miktari_entry.grid(row=1, column=1, padx=5, pady=5)

gider_para_birimi = ttk.Combobox(gider_frame, values=["TL", "USD", "EUR"], font=("Arial", 10), state="readonly")
gider_para_birimi.current(0)  # Varsayılan olarak TL seçili
gider_para_birimi.grid(row=2, column=1, padx=5, pady=5)

gider_para_birimi_label = tk.Label(gider_frame, text="Para Birimi:", font=("Arial", 10))
gider_para_birimi_label.grid(row=2, column=0, padx=5, pady=5)

gider_tarihi_label = tk.Label(gider_frame, text="Tarih:", font=("Arial", 10))
gider_tarihi_label.grid(row=3, column=0, padx=5, pady=5)
gider_tarihi_entry = DateEntry(gider_frame, font=("Arial", 10), width=37, background='darkblue', foreground='white', borderwidth=2)
gider_tarihi_entry.grid(row=3, column=1, padx=5, pady=5)

gider_saat_combobox = ttk.Combobox(gider_frame, values=[f"{i:02}" for i in range(24)], font=("Arial", 10), state="readonly")
gider_saat_combobox.grid(row=4, column=1, padx=5, pady=5)
gider_saat_combobox.current(0)

gider_dakika_combobox = ttk.Combobox(gider_frame, values=[f"{i:02}" for i in range(60)], font=("Arial", 10), state="readonly")
gider_dakika_combobox.grid(row=5, column=1, padx=5, pady=5)
gider_dakika_combobox.current(0)

gider_saat_label = tk.Label(gider_frame, text="Saat:", font=("Arial", 10))
gider_saat_label.grid(row=4, column=0, padx=5, pady=5)

gider_dakika_label = tk.Label(gider_frame, text="Dakika:", font=("Arial", 10))
gider_dakika_label.grid(row=5, column=0, padx=5, pady=5)

tekrarla_var = tk.BooleanVar()
tekrarlayan_gider_checkbox = tk.Checkbutton(gider_frame, text="Tekrarlayan Gider", variable=tekrarla_var, font=("Arial", 10))
tekrarlayan_gider_checkbox.grid(row=6, column=0, columnspan=2, padx=5, pady=5)

gider_ekle_button = tk.Button(gider_frame, text="Gider Ekle", command=gider_ekle, font=("Arial", 10), bg="lightcoral")
gider_ekle_button.grid(row=7, columnspan=2, padx=5, pady=5, ipadx=10)

# Hesaplama ve kaydetme butonları
hesapla_button = tk.Button(root, text="Hesapla", command=hesapla, font=("Arial", 12), bg="lightblue")
hesapla_button.pack(pady=10)

sonuc_label = tk.Label(root, text="", font=("Arial", 14))
sonuc_label.pack(pady=10)

excel_kaydet_button = tk.Button(root, text="Excel'e Kaydet", command=excel_kaydet, font=("Arial", 12), bg="lightyellow")
excel_kaydet_button.pack(pady=10)

root.mainloop()
